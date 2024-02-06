import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# Print versions for debugging
print("Pandas version:", pd.__version__)
print("OpenPyXL version:", openpyxl.__version__)


def create_summary_table(file_path):
    print(f"Loading workbook from {file_path}")
    wb = load_workbook(file_path)
    if 'Working Copy' not in wb.sheetnames:
        first_sheet = wb.worksheets[0]
        first_sheet.title = 'Working Copy'
        print("Created 'Working Copy' sheet as it was not present.")
    ws = wb['Working Copy']

    bold_font = Font(bold=True)

    print("Extracting data into DataFrame...")
    data = ws.iter_rows(min_row=2, values_only=True)
    headers = next(data)  # Get the header row
    df = pd.DataFrame(data, columns=headers)
    print("Data loaded into DataFrame.")

    print("Grouping data and calculating sum of 'Net Bookings'...")
    summary_df = df.groupby(['Name', 'MFR Name']).agg({'Net Bookings': 'sum'}).reset_index()
    print("Grouping and summation complete.")

    if 'Table' in wb.sheetnames:
        del wb['Table']
        print("'Table' sheet existed and was deleted.")
    ws_summary = wb.create_sheet('Table')
    print("Created 'Table' sheet for summary.")

    ws_summary.append(['Customer Names By Region:', 'Sum of Net Bookings:'])
    blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    currency_format = '"$"#,##0.00'
    row_num = 2

    for name, group in summary_df.groupby('Name'):
        print(f"Processing: {name}")
        name_cell = ws_summary.cell(row=row_num, column=1, value=name)
        name_cell.fill = blue_fill
        name_cell.font = bold_font
        total_net_bookings_cell = ws_summary.cell(row=row_num, column=2, value=group['Net Bookings'].sum())
        total_net_bookings_cell.number_format = currency_format
        total_net_bookings_cell.font = bold_font
        row_num += 1

        for _, row in group.iterrows():
            ws_summary.cell(row=row_num, column=1, value='    ' + row['MFR Name'])
            net_bookings_cell = ws_summary.cell(row=row_num, column=2, value=row['Net Bookings'])
            net_bookings_cell.number_format = currency_format
            row_num += 1

    print("Calculating grand total...")
    grand_total = 0
    for row in ws_summary.iter_rows(min_row=2, max_col=2, max_row=ws_summary.max_row):
        if row[0].font.bold:
            grand_total += row[1].value

    grand_total_row = ws_summary.max_row + 1
    ws_summary.cell(row=grand_total_row, column=1, value="Grand Total").font = Font(bold=True)
    grand_total_cell = ws_summary.cell(row=grand_total_row, column=2, value=grand_total)
    grand_total_cell.number_format = currency_format
    grand_total_cell.font = Font(bold=True)
    print(f"Grand total calculated and written: {grand_total}")

    print("Adjusting column widths...")
    for col in ws_summary.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        adjusted_width = (max_length + 2) * 1.2
        ws_summary.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    print("Saving workbook...")
    wb.save(file_path)
    print("Workbook saved. Process complete.")
