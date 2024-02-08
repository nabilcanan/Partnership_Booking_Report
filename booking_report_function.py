from tkinter import filedialog
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from tkinter import messagebox
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from table import create_summary_table


# Sheet 1 needs to be 'Working Copy'
def on_file_selected():
    file_path = select_excel_file()
    if file_path:
        process_excel_file(file_path)


def select_excel_file():
    # Show an "Open" dialog box and return the path to the selected file
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:  # Check if a file was selected
        print(f"File selected: {file_path}")
        return file_path
    else:
        print("No file was selected.")
        return None


# Logic we call when we need to execute functions in order
def process_excel_file(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    shift_worksheet_down(ws)
    highlight_duplicate_values(ws)
    add_headers_and_formulas(ws)
    format_columns_as_currency_and_percentage(ws)
    format_specific_columns_as_text(ws, ['Qty', 'Ship&Debit'])
    highlight_header_cells(ws)
    freeze_panes(ws)
    format_net_bookings_column(ws)
    add_total_net_bookings(ws)  # Call the new function to add total and highlight
    save_and_notify(file_path, wb)
    create_summary_table(file_path)


# ---------- Add Net Booking Logic -------------------------------------------------------------
def add_total_net_bookings(ws):
    net_bookings_col_index = None

    for col in ws.iter_cols(min_row=2, max_row=2):
        for cell in col:
            if cell.value == 'Net Bookings':
                net_bookings_col_index = cell.column
                break
        if net_bookings_col_index:
            break

    if net_bookings_col_index:
        total_formula = f'=SUM({get_column_letter(net_bookings_col_index)}3:{get_column_letter(net_bookings_col_index)}{ws.max_row})'
        # Check if the style already exists, create a unique name if it does
        style_name = "currency"
        while style_name in ws.parent.named_styles:
            style_name += "_duplicate"
        currency_style = NamedStyle(name=style_name, number_format='"$"#,##0.00')
        ws.parent.add_named_style(currency_style)
        ws.cell(row=1, column=net_bookings_col_index, value=total_formula)
        ws.cell(row=1, column=net_bookings_col_index).style = currency_style
        # Apply the light yellow fill
        light_yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        ws.cell(row=1, column=net_bookings_col_index).fill = light_yellow_fill


# ---------- End of Add Net Booking Logic -------------------------------------------------------------


def shift_worksheet_down(ws):
    ws.insert_rows(1)


# ------------------- Highlight Duplicate Values Func -------------------------------------
def highlight_duplicate_values(ws):
    seen_values = {}
    lighter_red_fill = PatternFill(start_color='FFDDDD', end_color='FFDDDD', fill_type='solid')

    for row in ws.iter_rows(min_col=1, max_col=1, min_row=2):
        cell = row[0]
        value = cell.value
        value = value.lstrip('0') if isinstance(value, str) else value
        cell.value = value
        if value in seen_values:
            seen_values[value].append(cell.coordinate)
        else:
            seen_values[value] = [cell.coordinate]

    for duplicates in seen_values.values():
        if len(duplicates) > 1:
            for coordinate in duplicates:
                ws[coordinate].fill = lighter_red_fill


# -------------------  End Of Highlight Duplicate Values Func ----------------------------------

# ----------------- Adding Headers to desired columns -----------------------------------------
def add_headers_and_formulas(ws):

    ws.insert_cols(13, 2)  # Need to add this line before adding new column in existing dataframes
    # This will ensure the existing columns are overwritten

    ws.cell(row=2, column=13, value='Total Cost')
    ws.cell(row=2, column=14, value='GP%')

    for row in range(3, ws.max_row + 1):
        cost_formula = f'=K{row}-L{row}'
        gp_percentage_formula = f'=(J{row}-M{row})/J{row}'
        ws[f'M{row}'] = cost_formula
        ws[f'N{row}'] = gp_percentage_formula

    ws.insert_cols(23, 4)
    ws.cell(row=2, column=23, value='Notes')
    ws.cell(row=2, column=24, value='Actual Cost')
    ws.cell(row=2, column=25, value='Resale Price')
    ws.cell(row=2, column=26, value='Actual GP%')

    for row in range(3, ws.max_row + 1):
        gp_percentage_formula = f'=(Y{row}-X{row})/Y{row}'
        ws[f'Z{row}'] = gp_percentage_formula
        ws[f'Y{row}'] = ws[f'I{row}'].value


# ----------------- End Of Adding Headers to desired columns ------------------------------------

# ---------------- Formatting Currency and Percentages for certain columns ---------------------
def format_columns_as_currency_and_percentage(ws):
    # Ensure the currency style is defined
    currency_style_name = 'currency'
    if currency_style_name not in ws.parent.named_styles:
        currency_style = NamedStyle(name=currency_style_name, number_format='"$"#,##0.0000')
        ws.parent.add_named_style(currency_style)
    else:
        currency_style = ws.parent.named_styles[currency_style_name]

    # Specify the columns to format as currency, including column L
    currency_columns = ['M', 'N', 'Z', 'Y', 'K', 'L', 'G', 'H', 'I', 'X']

    # Apply the currency style to the cells in the specified columns
    for col_letter in currency_columns:
        for row in range(2, ws.max_row + 1):
            ws[col_letter + str(row)].style = currency_style

    # Create or get the percentage style
    percentage_style_name = 'percentage'
    if percentage_style_name not in ws.parent.named_styles:
        percentage_style = NamedStyle(name=percentage_style_name, number_format=FORMAT_PERCENTAGE_00)
        ws.parent.add_named_style(percentage_style)
    else:
        percentage_style = ws.parent.named_styles[percentage_style_name]

    # Columns to format as percentage
    percentage_columns = ['N', 'Z']

    # Apply the percentage style to the cells in the specified columns
    for col_letter in percentage_columns:
        for row in range(2, ws.max_row + 1):
            cell = ws[col_letter + str(row)]
            cell.style = percentage_style

# ---------------- End of Formatting Currency and Percentages for certain columns ----------------

# ---------------- Formatting Specific columns as text --------------------------------

def format_specific_columns_as_text(ws, text_columns):
    for col_name in text_columns:
        col_index = None

        for col in ws.iter_cols(min_row=2, max_row=2):
            for cell in col:
                if cell.value == col_name:
                    col_index = cell.column
                    break
            if col_index:
                break

        if col_index:
            for row in range(3, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_index)
                cell.number_format = '@'  # Format as text


# ---------------- End of Formatting Specific columns as text ---------------------------

# ---------------- Highlight header cells accordingly ------------------------------------
def highlight_header_cells(ws):
    header_cells = ['M2', 'N2', 'W2', 'X2', 'Y2', 'Z2']
    light_yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')

    for cell_reference in header_cells:
        ws[cell_reference].fill = light_yellow_fill


# ---------------- End of Highlight header cells accordingly -------------------------------

def freeze_panes(ws):
    ws.freeze_panes = 'E2'


def format_net_bookings_column(ws):
    net_bookings_col_index = None

    for col in ws.iter_cols(min_row=2, max_row=2):
        for cell in col:
            if cell.value == 'Net Bookings':
                net_bookings_col_index = cell.column
                break
        if net_bookings_col_index:
            break

    if net_bookings_col_index:
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row=row, column=net_bookings_col_index)
            cell.number_format = '"$"#,##0.00'
    else:
        print("Column 'Net Bookings' not found")


def save_and_notify(file_path, wb):
    wb.save(file_path)
    messagebox.showinfo("Processing Complete", "Excel file processing is complete. You can now save the file.")
